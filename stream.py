# 页面访问冒烟测试
import streamlit as st
import requests
import openpyxl
import wget
import pandas as pd
import os
import base64



# 设置页面标题
st.set_page_config(page_title ="接口测试自动化平台",)


# 这里是页面
import streamlit as st
from PIL import Image
st.markdown(
 '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@4.5.3/dist/css/bootstrap.min.css" integrity="sha384-TX8t27EcRE3e/ihU7zmQxVncDAy5uIKz4rEkgIXeMed4M0jlfIDPvg6uqKI2xXr2" crossorigin="anonymous">',
 unsafe_allow_html=True,
)
query_params = st.experimental_get_query_params()
# query_params123 = st.query_params()

tabs = ["首页", "图片", "接口管理", "视频"]
if "tab" in query_params:
    active_tab = query_params["tab"][0]
else:
    active_tab = "首页"

if active_tab not in tabs:
    st.query_params(tab="首页")
    active_tab = "首页"

li_items = "".join(
    f"""
    <li class="nav-item">
        <a class="nav-link{' active' if t==active_tab else ''}" href="/?tab={t}">{t}</a>
    </li>
    """
    for t in tabs
)
tabs_html = f"""
    <ul class="nav nav-tabs">
    {li_items}
    </ul>
"""

st.markdown(tabs_html, unsafe_allow_html=True)
st.markdown("<br>", unsafe_allow_html=True)

if active_tab == "首页":
    st.info("中国人自己的空间站时代来临了")
    image1 = Image.open('./中国空间站.png')
    st.image(image1, use_column_width='auto', caption='中国空间站')
elif active_tab == "图片":
    st.info("请欣赏雪山景色")
    image = Image.open('./图片/图片 (5).jpg')
    st.image(image, use_column_width='auto', caption='雪山景色',output_format="png")
elif active_tab == "接口管理":
    st.info('点击播放按钮接口管理')
    audio_file = open('./视频与图像/Welcome.mp3', 'rb')
    audio_bytes = audio_file.read()
    st.audio(audio_bytes, format='audio/mp3')
elif active_tab == "视频":
    st.info('点击播放按钮播放mp4视频')
    video_file = open('./视频与图像/star.mp4', 'rb')
    video_bytes = video_file.read()
    st.video(video_bytes)
else:
    st.error("出错了。")





from businesstest import ceshi



st.title('权和益c端接口自动化测试平台')

class apitest():

    def write_result(self,filename,sheetname,row,column,final_result):

        wb=openpyxl.load_workbook(filename) #加载文件
        sheet=wb[sheetname] #读取sheet
        sheet.cell(row=row,column=column).value=final_result #将结果写入cell
        wb.save(filename) #保存文件

    def read_data(self,filename,sheetname):
        wb=openpyxl.load_workbook(filename) #加载工作薄
        sheet=wb[sheetname] #获取表单
        max_row=sheet.max_row #获取最大行数
        case_list=[] #创建一个空列表，存放测试用例

        for i in range(2,max_row+1):
            dict1=dict(
            case_id=sheet.cell(row=i,column=1).value, #获取case_id
            name=sheet.cell(row=i,column=2).value,  #获取name 接口名字
            url=sheet.cell(row=i,column=3).value,   #获取url
            method=sheet.cell(row=i,column=4).value,   #获取请求方式
            headers=sheet.cell(row=i,column=5).value, #获取请求头
            data=sheet.cell(row=i,column=6).value,  #获取请求数据data

            response=sheet.cell(row=i,column=7).value #获取期望结果
            )
            case_list.append(dict1)  #每循环一次就将一行数据加入列表
        return case_list  #返回测试用例列表

    def duanyan(self,response,msg,case_id,url):
            bb={}
            res = response.json()
            print(res)
            real_msg=res.get("msg")

            print("预期结果{}".format(msg))
            print("实际结果{}".format((real_msg)))


            if real_msg==msg:
                print("第{}条用例通过".format(case_id))
                final_re="passed"
                mm="{}条".format(case_id)+"{}接口请求成功".format(url)

            else:
                print("第{}条用例不通过".format(case_id))
                final_re="failed"
                mm = "{}条".format(case_id)+"{}接口请求失败".format(url)+real_msg
            # 写入用例通过结果
            bb['final_re']=final_re
            bb['mm']=mm
            return bb

    def testreport(self,databiao):
        biao={"main":"main_api.xlsx","test":"test_api.xlsx"}
        dataxlsx=biao[databiao]
        print(dataxlsx)
        data = open(dataxlsx, 'rb').read()  # 以只读模式读取且读取为二进制文件

        b64 = base64.b64encode(data).decode('UTF-8')  # 解码并加密为base64
        href = f'<a href="data:file/data;base64,{b64}" download="myresults.xlsx">下载测试报告</a>'  # 定义下载链接，默认的下载文件名是myresults.xlsx
        st.markdown(href, unsafe_allow_html=True)  # 输出到浏览器

    def api_test(self,cases): #传入读取到的测试用例数据
        print(cases)
        kk=[]
        for case in cases:
            print(case)


            case_id=case.get("case_id") #从列表中取出case_id
            url=case.get("url")#从列表中取出url
            method=case.get("method")#从列表中取出method
            headers=eval(case.get("headers"))#从列表中取出headers
            # return headers
            data=eval(case.get("data"))#从列表中取出data注意取出来的是一个字符串，用eval（）函数可以去掉外面的双引号，直接用内部的值；
            response=eval(case.get("response"))#从列表中取出response

            msg=response.get("msg")#从response中取出msg

            if method=="get": #判断请求方式
                response=requests.get(url=url,headers=headers,data=data)

            else:   #这是post请求
                response=requests.post(url=url,headers=headers,json=data,verify=False)#调用请求方法，传入从列表取出的url，data,hedaers
            bb = self.duanyan(response, msg, case_id, url) #调用断言方法返回通过结果列表
            self.write_result("test_api.xlsx", "Sheet1", case_id+1, 8, bb['final_re'])#调用写入excel函数
            baogao=bb['mm']
            kk.append(baogao)

        return kk

    # 支付功能
    def mainpay(self,cases):

        url = "https://yp.qhypay.cn/retail/user/wallet/pay_wallet_info"
        headers = {"Content-Type": "application/json"}
        data = {
            "token": "MzowOGY5b1JKZVMvMGQ2Ym1MMEJNbnNjb1BseHpJWHl3UVlPOS9Idi9UYUpHeDJYWW83eW9KOjE3MDgzMzAwNzE=",
            "version": "2.6.2",
            "plat": "android",
            "plat_ver": "QP1A.190711.020 release-keys",
            "app_plat": "app",
            "app_ver": "1.0.0.2",
            "mobile_model": "V1838A",
            "client": "app",
            "device_id": "e56f17d1-0430-488b-ae20-9cab485621d7",
            "getui_cid": "5394b115a1b684cb6a21fe394b19d430",
            "deviceToken": "",
            "timestamp": 1707034131,
            "sign": "6c421c1e798f1e24218554f778195c6d"}
        for case in cases:
            print(case)
   # 先获取账号余额和积分


            case_id=case.get("case_id") #从列表中取出case_id
            url=case.get("url")#从列表中取出url
            method=case.get("method")#从列表中取出method
            headers=eval(case.get("headers"))#从列表中取出headers
            # return headers
            data=eval(case.get("data"))#从列表中取出data注意取出来的是一个字符串，用eval（）函数可以去掉外面的双引号，直接用内部的值；
            response=eval(case.get("response"))#从列表中取出response
            msg = response.get("msg")  # 从response中取出msg
            res=requests.post(url=url,json=data,headers=headers,verify=False).json()
            bb = self.duanyan(res, msg, case_id, url)

# 支付信息

# 测试环境接口测试
if st.button('测试环境接口测试') :
    st.write('=============开始测试=============')
    aa=apitest()
    cases=aa.read_data("test_api.xlsx","Sheet1")
    # print(cases)
    pp=aa.api_test(cases)
    st.text(pp)
    for p in pp:
        st.text(p)
    st.write('==========测试结束=========')
    aa.testreport('test')

#正式环境接口测试
if st.button('正式环境接口测试') :
    st.write('============测试开始==============')
    aa=apitest()
    cases=aa.read_data("main_api.xlsx","Sheet1")
    pp = aa.api_test(cases)
    st.text(pp)
    for p in pp:
        st.text(p)
    st.write('==========测试结束=========')
    aa.testreport()

# col1, col2, col3 = st.columns([1,1,1])
# with col1:
#     # st.button('测试环境接口测试')
#
# with col2:
#     st.button('正式环境接口测试2')
# with col3:
#     st.button('3')

if st.button("测试环境测试支付功能") :
    st.write('Why hello there')
    aa=apitest()
    cases=aa.read_data("main_api.xlsx","Sheet1")
    pp = aa.api_test(cases)
    st.text(pp)
    for p in pp:
        st.text(p)

if st.button("正式环境测试支付功能") :

    st.write('Why hello there')
    aa = apitest()
    aa.mainpay()
    # cases = aa.read_data("main_api.xlsx", "Sheet1")
    # pp = aa.api_test(cases)
    # st.text(pp)
    # for p in pp:
    #     st.text(p)









print(os.path.abspath(__file__))





