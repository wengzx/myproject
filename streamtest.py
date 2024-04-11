#-*- coding : utf-8-*-


import json
import time

import streamlit as st
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import  PatternFill  

import pandas as pd

import base64

import schedule   #定时任务

import yagmail



# 设置页面标题
st.set_page_config(
        page_title="自动化接口测试平台",
)

# 定义多页面
class MultiApp:

    def __init__(self):
        self.apps = []
        self.app_dict = {}


    def add_app(self, title, func):
        if title not in self.apps:
            self.apps.append(title)
            self.app_dict[title] = func


    def run(self):
        title = st.sidebar.radio(
            # 'Go To',
            '接口测试',
            self.apps,
            format_func=lambda title: str(title))
        self.app_dict[title]()


# api测试方法
class apitest():

    # 写入excel文件结果

    def write_result(self,filename,sheetname,row,column,final_result):

        #Color=['c6efce','006100']#绿
        #Color = ['ffc7ce', '9c0006']  #红
        #Color = ['ffeb9c', '9c6500']  # 黄
        Color=['75BD42','FF0000']
        if final_result=='passed':
            final_result_color= PatternFill('solid',Color[0])
        else:
            final_result_color= PatternFill('solid',Color[1])
        wb=openpyxl.load_workbook(filename) #加载文件
        sheet=wb[sheetname] #读取sheet
         #sheet.cell(row=row,column=column).value=final_result
        sheet.cell(row=row,column=column,value=final_result).fill=final_result_color #将结果写入cell
        wb.save(filename) #保存文件


    # 读取excel文件

    def read_data(_self,filename,sheetname):
        wb=openpyxl.load_workbook(filename) #加载工作薄
        sheet=wb[sheetname] #获取表单
        max_row=sheet.max_row #获取最大行数
        case_list=[] #创建一个空列表，存放测试用例

        # 循环
        for i in range(2,max_row+1):
            dict1=dict(
            case_id=sheet.cell(row=i,column=1).value, #获取case_id
            name=sheet.cell(row=i,column=2).value,  #获取name 接口名字
            url=sheet.cell(row=i,column=3).value,   #获取url
            method=sheet.cell(row=i,column=4).value,   #获取请求方式
            headers=sheet.cell(row=i,column=5).value, #获取请求头
            data=sheet.cell(row=i,column=6).value,  #获取请求数据data

            response=sheet.cell(row=i,column=7).value #获取期望结果
            )
            case_list.append(dict1)  #每循环一次就将一行数据加入列表
        return case_list  #返回测试用例列表


    # 断言结果是否与excel预期结果一致

    def duanyan(self,response,msg,case_id,url):
            bb={}  #定义一个空字典
            print(type(response))

            if type(response)!=dict:  #如果返回数据类型不是字典

                res = response.json()  #json返回数据，将其json化成python可读数据
            else:
                print(55)
            # print(res)
                res=response

                print(res)
                real_msg=res.get("msg")   #获取返回数据的msg
                print(real_msg)
                # sys.exit()

                print("预期结果{}".format(msg))
                print("实际结果{}".format((real_msg)))


                if real_msg==msg:             #如果返回msg和预期结果msg一致
                    print("第{}条用例通过".format(case_id))
                    final_re="passed"
                    mm="{}条".format(case_id)+"{}接口请求成功".format(url)

                else:                      ##如果返回msg和预期结果msg一致
                    print("第{}条用例不通过".format(case_id))
                    final_re="failed"
                    mm = "{}条".format(case_id)+"{}接口请求失败".format(url)+real_msg
                # 写入用例通过结果
                bb['final_re']=final_re
                bb['mm']=mm
                return bb


    # 生成测试报告

    def testreport(self,databiao):   #传入databiao
        biao={"main":"main_api.xlsx","test":"test_api.xlsx","pay":"pay_api.xlsx"}   #将文件用字典的键值存储起来
        dataxlsx=biao[databiao]    #获取excel文件名
        print(dataxlsx)
        now=int(time.time())     #获取时间
        data = open(dataxlsx, 'rb').read()  # 以只读模式读取且读取为二进制文件

        b64 = base64.b64encode(data).decode('UTF-8')  # 解码并加密为base64
        href = f'<a href="data:file/data;base64,{b64}" download="{now}_myresults.xlsx">下载测试报告</a>'  # 定义下载链接，默认的下载文件名是myresults.xlsx
        st.markdown(href, unsafe_allow_html=True)  # 输出到浏览器


    # 测试所有接口

    def api_test(self,cases): #传入读取到的测试用例数据
        self.token=''       #定义token
        # print(cases)
        kk=[]
        for case in cases:     #循环每一行excel数据在读取的excel文件内容
            err=""
            if case['case_id'] is None :     #如果excel的case_id列内容为空
                baogao="表格第{}条".format(case_id+1) + "内容为空,请删除此行"
                print(baogao)
                continue
            print(case)
            print(type(case))

            name=case.get("name") #从列表中取出name
            case_id=case.get("case_id") #从列表中取出case_id
            url=case.get("url")#从列表中取出url
            method=case.get("method")#从列表中取出method
            headers=eval(case.get("headers"))#从列表中取出headers
            print(headers)
            if 'token' in headers :    #如果请求头包含token
                print(2223)
                headers['token']=self.token  #将全局变量token赋值请求头的token
            else:
               pass

            print("data数据")
            print(type(case["data"]))
            print(case["data"])

            try:
                data=eval(case.get("data"))   #获取data的参数的值
            except Exception as e:
                print(e)
                err="请检查data参数里面是否有值未加双引号"
            print(333)
            # data=eval(case.get("data"))#从列表中取出data注意取出来的是一个字符串，用eval（）函数可以去掉外面的双引号，直接用内部的值；
            # 将旧的token更换为新token
            data['token']=self.token
            print( "tokne"+data['token'])

            response=eval(case.get("response"))#从列表中取出response

            msg=response.get("msg")#从response中取出msg
            print(case['case_id'])

            if method=="get": #判断请求方式
                response=requests.get(url=url,headers=headers,data=data)

            else:   #这是post请求
                res=requests.post(url=url,headers=headers,json=data,verify=False)#调用请求方法，传入从列表取出的url，data,hedaers
                response = res.json()   #返回数据json化成python可读的数据
                aa=type(response)
                print(aa)


                rstoken=response["data"]
                print(rstoken)
                if 'token' in rstoken:      #如果返回数据有token
                    print(rstoken['token'])

                    tokens=rstoken['token']
                    self.token = tokens      #将token的值赋值给全职变量self.token
                    print(self.token)



            bb = self.duanyan(response, msg, case_id, url) #调用断言方法返回通过结果列表
            self.write_result("test_api.xlsx", "Sheet1", case_id+1, 8, bb['final_re'])#调用写入excel函数
            baogao=bb['mm']+err
            print(baogao)
            if "支付" in name:    #如果接口名字有支付两个字
                kk.append(name)
                self.write_result("pay_api.xlsx", "Sheet1", case_id + 1, 8, bb['final_re'])  #写入支付接口excel文件
            kk.append(baogao)

        return kk   #返回内容


    def  send_email(self):
        # 链接邮箱服务器  password授权码
        yag = yagmail.SMTP(user="1294424625@qq.com", password="dmaosvisrjzaibai", host='smtp.qq.com')

        # 邮箱正文
        # contents = ['This is the body, and here is just text http://somedomain/image.png',
        #             'You can find an audio file attached.', '/local/path/song.mp3']

        contents = ['接口测试有请求失败接口']

        # 发送邮件
        yag.send('1294424625@qq.com', 'subject', contents)


    # 定时任务任务函数具体方法

    def task_func(self):
        print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))

        # return schedule.CancelJob
        st.write('=============开始定时任务=============')
        aa = apitest()
        cases = aa.read_data("test_api.xlsx", "Sheet1")
        # print(cases)
        pp = aa.api_test(cases)
        # st.text(pp)
        for p in pp:
            if "失败" in p:
                st.text(p)
                aa.send_email()


    # 调用任务函数主方法
    def task(self,name):
        if name == 2:

            schedule.clear()
        # schedule.every(50).seconds.do(self.task_func)
        schedule.every().day.at("09:30").do(self.task_func)     #每天11点半执行定时任务


        while True:
            schedule.run_pending()








    # def send_email(self):






# 页面背景设置

def main_bg(main_bg):
    main_bg_ext = "png"
    st.markdown(
        f"""
         <style>
         .stApp {{
             background: url(data:image/{main_bg_ext};base64,{base64.b64encode(open(main_bg, "rb").read()).decode()});
             background-size: cover
         }}
         </style>
         """,
        unsafe_allow_html=True
    )


# 调用
# main_bg('./1111.jpg')


# 侧边栏背景设置

def sidebar_bg(side_bg):
    side_bg_ext = 'png'

    st.markdown(
        f"""
      <style>
      [data-testid="stSidebar"] > div:first-child {{
          background: url(data:image/{side_bg_ext};base64,{base64.b64encode(open(side_bg, "rb").read()).decode()});
      }}
      </style>
      """,
        unsafe_allow_html=True,
    )


# 调用
# sidebar_bg('./999.jpg')



def make_clickable(url):
    return f'<a target="_blank" href="{url}">www.baidu.com</a>'



def convert_df(df):
    # return df.to_csv("1.csv")
    return df.to_csv().encode('utf-8')


# 页面具体自定义内容

def foo():
    # st.title("请开始你的测试吧！！！")


    if st.button('测试环境接口测试'):
        st.write('=============开始测试=============')
        aa = apitest()
        cases = aa.read_data("test_api.xlsx", "Sheet1")
        # print(cases)
        pp = aa.api_test(cases)
        st.text(pp)
        for p in pp:
            st.text(p)
        st.write('==========测试结束=========')
        aa.testreport('test')

    if st.button('测试环境支付功能测试'):
        st.write('=============开始测试=============')
        aa = apitest()
        cases = aa.read_data("pay_api.xlsx", "Sheet1")
        # print(cases)
        pp = aa.api_test(cases)
        st.text(pp)
        for p in pp:
            st.text(p)
            if "失败" in p:
                print(222)
                # st.markdown(":red[{p}]",p)
        st.write('==========测试结束=========')
        aa.testreport('pay')

        # 定时任务按钮




def bar():
    # st.title("请开始你的测试吧！！！")
    if st.button('正式环境接口测试'):
        st.write('============测试开始==============')
        aa = apitest()
        cases = aa.read_data("main_api.xlsx", "Sheet1")
        pp = aa.api_test(cases)
        st.text(pp)
        for p in pp:
            st.text(p)
        st.write('==========测试结束=========')
        aa.testreport()


def gen():
    st.title("接口管理")
    df1 = pd.read_excel('excel1.xlsx')
    data6 = df1.iloc[:, [0]].values
    print(data6)
    data = {
        # 'url': ['https://www.amazon.com/' for i in range(10)]
        'url': [data6]
    }
     # 读取指定列的所有行数据：读取第一列所有数据
    # df = pd.DataFrame(data)
    # df['url'] = df['url'].apply(make_clickable)
    # print(df)
    # df = df.to_html(escape=False)
    # st.write(df, unsafe_allow_html=True)
    aa="test_api.xlsx"
    df=pd.read_excel(aa)
    pp={}
    pp['name']=df['name']
    pp['url']=df['url']

    st.table(pp)

    # name = st.text_input('请输入用户名', max_chars=100, help='最大长度为100字符')


# 自定义接口管理页面调用方法

def diy():

    aa = apitest()
    aa.testreport('test')

    # st.markdown(get_binary_file_download_html('1.xlsx',"表格"),unsafe_allow_html=True)
    #
    st.write("<h5 style='color: red;'>请先下载测试报告模板，自定义接口数据再上传文件进行测试</h5>", unsafe_allow_html=True)

    up_file = st.file_uploader("上传文件", type=["xls", "xlsx", "csv"])
    print(up_file)
    if up_file is not None:
        print(222)
        head, sep, tail = str(up_file.name).partition(".")
        st.write("文件名称是：" + str(head))
        st.write("文件类型是：" + str(tail))
    #        sep是上传文件的文件名何文件类型的分隔符
        if tail=="xls" or tail=="xlsx":
            # df=pd.read_excel(up_file)
            # st.table(df)
            aa = apitest()
            cases = aa.read_data("test_api.xlsx", "Sheet1")
            # print(cases)
            pp = aa.api_test(cases)
            st.text(pp)
            for p in pp:
                st.text(p)
            st.write('==========测试结束=========')
            aa.testreport('test')

        elif tail=="csv":
            df=pd.read_csv(up_file)
            st.table(df)
    else:
        print(222)

def time_task():
    aa = apitest()
    if st.button("停止定时任务"):
        st.write("已经停止定时任务")
        aa.task(name=2)

    if st.button("开启定时任务(每天11点半执行定时任务)"):
        aa.task(name=1)







app = MultiApp()
app.add_app("测试环境", foo)
app.add_app("正式环境", bar)
app.add_app("接口管理", gen)
app.add_app("自定义接口数据测试", diy)
app.add_app("定时任务", time_task)


app.run()
