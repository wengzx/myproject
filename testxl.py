import  openpyxl
import time


bk=openpyxl.load_workbook('userinfo.xlsx') #打开一个文件
sheet=bk.active #打开工作表
# sheet1=bk.get_sheet_by_name('Sheet')  #打开工作表
minrow=sheet.min_row  #最小行'
maxrow=sheet.max_row  #最小行
mincol=sheet.min_column #最小列
maxcol=sheet.max_column


for i in range(minrow,maxrow+1):
    for j in range(mincol,maxcol+1):
        cell=sheet.cell(i,j).value
        print(cell,end="")
    print()



timestamp = time.time()
print("当前时间戳：", int(timestamp))